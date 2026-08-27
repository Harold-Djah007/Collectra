from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from django.test import RequestFactory, SimpleTestCase
from lxml import etree
from openpyxl import Workbook

from corehq.apps.app_manager.xlsform import (
    JAVAROSA_NAMESPACE,
    XHTML_NAMESPACE,
    XFORMS_NAMESPACE,
    XlsFormDefinition,
    XlsFormError,
    build_xform,
    parse_xlsform,
)
from corehq.apps.app_manager.views.form_builder import NEW_MODULE_VALUE, _save_draft


def _workbook_file(survey_rows, choice_rows=None, setting_rows=None):
    workbook = Workbook()
    survey = workbook.active
    survey.title = "survey"
    for row in survey_rows:
        survey.append(row)
    if choice_rows:
        choices = workbook.create_sheet("choices")
        for row in choice_rows:
            choices.append(row)
    if setting_rows:
        settings = workbook.create_sheet("settings")
        for row in setting_rows:
            settings.append(row)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _representative_xlsform():
    return _workbook_file(
        [
            [
                "type", "name", "label::English (en)", "label::French (fr)",
                "hint::English (en)", "required", "relevant", "constraint",
                "constraint_message::English (en)", "calculation", "choice_filter",
            ],
            ["start", "start", "", "", "", "", "", "", "", "", ""],
            ["begin group", "identity", "Identity", "Identité", "", "", "", "", "", "", ""],
            ["text", "full_name", "Full name", "Nom complet", "As on ID", "yes", "", "", "", "", ""],
            ["integer", "age", "Age", "Âge", "", "yes", "", ". >= 18", "Must be 18 or older", "", ""],
            ["select_one regions", "region", "Region", "Région", "", "", "", "", "", "", ""],
            [
                "select_one districts", "district", "District", "District", "", "", "${region} != ''",
                "", "", "", "${region} = region",
            ],
            ["begin repeat", "visits", "Visits", "Visites", "", "", "", "", "", "", ""],
            ["date", "visit_date", "Visit date", "Date de visite", "", "", "", "", "", "", ""],
            ["end repeat", "", "", "", "", "", "", "", "", "", ""],
            ["note", "consent_note", "Ask for consent", "Demander le consentement", "", "", "", "", "", "", ""],
            ["end group", "", "", "", "", "", "", "", "", "", ""],
        ],
        [
            ["list_name", "name", "label::English (en)", "label::French (fr)", "region"],
            ["regions", "north", "North", "Nord", ""],
            ["regions", "south", "South", "Sud", ""],
            ["districts", "north_one", "North One", "Nord Un", "north"],
            ["districts", "south_one", "South One", "Sud Un", "south"],
        ],
        [
            ["form_title", "form_id", "default_language", "version", "instance_name"],
            ["Household Survey", "household_survey", "English (en)", "2026.1", "Household"],
        ],
    )


class XlsFormParserTest(SimpleTestCase):
    def test_parses_surveycto_structure_and_translations(self):
        definition = parse_xlsform(_representative_xlsform(), "household.xlsx")

        self.assertEqual(definition.form_title, "Household Survey")
        self.assertEqual(definition.form_id, "household_survey")
        self.assertEqual(definition.default_language, "en")
        self.assertEqual(definition.languages, ["en", "fr"])
        self.assertEqual(definition.question_count, 6)
        self.assertFalse(definition.errors)
        self.assertEqual(definition.rows[2].required, "true()")
        self.assertEqual(definition.rows[3].constraint, ". >= 18")

    def test_definition_round_trips_through_cache_safe_dictionary(self):
        definition = parse_xlsform(_representative_xlsform(), "household.xlsx")

        restored = XlsFormDefinition.from_dict(definition.to_dict())

        self.assertEqual(restored.to_dict(), definition.to_dict())
        self.assertEqual(restored.question_count, definition.question_count)

    def test_reports_missing_choice_list_without_saving(self):
        stream = _workbook_file([
            ["type", "name", "label"],
            ["select_one missing", "answer", "Answer"],
        ])

        definition = parse_xlsform(stream, "missing-choices.xlsx")

        self.assertEqual(len(definition.errors), 1)
        self.assertIn("Choice list 'missing'", definition.errors[0].message)

    def test_preserves_underscores_in_choice_list_names(self):
        stream = _workbook_file(
            [
                ["type", "name", "label"],
                ["select_one customer_type_list", "customer_type", "Customer type"],
            ],
            [
                ["list_name", "name", "label"],
                ["customer_type_list", "retail", "Retail customer"],
            ],
        )

        definition = parse_xlsform(stream, "customer-types.xlsx")

        self.assertFalse(definition.errors)
        self.assertEqual(definition.rows[0].list_name, "customer_type_list")

    def test_reports_duplicate_names_and_unclosed_groups(self):
        stream = _workbook_file([
            ["type", "name", "label"],
            ["begin group", "details", "Details"],
            ["text", "answer", "First"],
            ["text", "answer", "Second"],
        ])

        definition = parse_xlsform(stream, "invalid.xlsx")

        messages = [issue.message for issue in definition.errors]
        self.assertTrue(any("already used" in message for message in messages))
        self.assertTrue(any("missing 'end group'" in message for message in messages))

    def test_rejects_workbook_without_survey_sheet(self):
        workbook = Workbook()
        workbook.active.title = "questions"
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        with self.assertRaisesMessage(XlsFormError, "sheet named 'survey'"):
            parse_xlsform(stream, "not-xlsform.xlsx")


class XlsFormCompilerTest(SimpleTestCase):
    def setUp(self):
        self.definition = parse_xlsform(_representative_xlsform(), "household.xlsx")
        self.xml = build_xform(self.definition)
        self.document = etree.fromstring(self.xml.encode())
        self.namespaces = {
            "h": XHTML_NAMESPACE,
            "xf": XFORMS_NAMESPACE,
            "jr": JAVAROSA_NAMESPACE,
        }

    def test_compiles_nested_data_binds_and_repeat(self):
        self.assertEqual(
            len(self.document.xpath("//xf:bind[@nodeset='/data/identity/age']", namespaces=self.namespaces)),
            1,
        )
        age_bind = self.document.xpath(
            "//xf:bind[@nodeset='/data/identity/age']", namespaces=self.namespaces,
        )[0]
        self.assertEqual(age_bind.get("type"), "xsd:int")
        self.assertEqual(age_bind.get("required"), "true()")
        self.assertEqual(age_bind.get("constraint"), ". >= 18")
        self.assertEqual(
            len(
                self.document.xpath(
                    "//xf:repeat[@nodeset='/data/identity/visits']", namespaces=self.namespaces,
                )
            ),
            1,
        )
        repeat_node = self.document.xpath(
            "//xf:instance[not(@id)]/*//*[local-name()='visits']", namespaces=self.namespaces,
        )[0]
        self.assertEqual(repeat_node.get(f"{{{JAVAROSA_NAMESPACE}}}template"), "")

    def test_compiles_translations_and_static_choices(self):
        translations = self.document.xpath("//xf:translation", namespaces=self.namespaces)
        self.assertEqual([item.get("lang") for item in translations], ["en", "fr"])
        self.assertTrue(
            self.document.xpath(
                "//xf:text[@id='identity/full_name-label']/xf:value[text()='Full name']",
                namespaces=self.namespaces,
            )
        )
        self.assertTrue(
            self.document.xpath("//xf:item/xf:value[text()='north']", namespaces=self.namespaces)
        )

    def test_compiles_filtered_choice_instance_and_references(self):
        district_bind = self.document.xpath(
            "//xf:bind[@nodeset='/data/identity/district']", namespaces=self.namespaces,
        )[0]
        self.assertEqual(district_bind.get("relevant"), "/data/identity/region != ''")
        itemset = self.document.xpath(
            "//xf:select1[@ref='/data/identity/district']/xf:itemset", namespaces=self.namespaces,
        )[0]
        self.assertIn("/data/identity/region = region", itemset.get("nodeset"))
        self.assertEqual(itemset.xpath("xf:label", namespaces=self.namespaces)[0].get("ref"), "jr:itext(label)")
        choice_instance = self.document.xpath(
            "//xf:instance[@id='choices_district']", namespaces=self.namespaces,
        )[0]
        self.assertEqual(choice_instance[0].tag, "root")
        model_children = list(choice_instance.getparent())
        first_bind = self.document.xpath("//xf:model/xf:bind", namespaces=self.namespaces)[0]
        self.assertLess(model_children.index(choice_instance), model_children.index(first_bind))

    def test_preserves_form_id_as_submission_namespace(self):
        data = self.document.xpath("//xf:instance[not(@id)]/*", namespaces=self.namespaces)[0]

        self.assertEqual(
            etree.QName(data).namespace,
            "http://openrosa.org/formdesigner/household_survey",
        )


class XlsFormDraftSaveTest(SimpleTestCase):
    @patch("corehq.apps.app_manager.views.form_builder.messages.success")
    @patch("corehq.apps.app_manager.views.form_builder.clear_app_cache")
    @patch("corehq.apps.app_manager.views.form_builder.cache.delete")
    @patch("corehq.apps.app_manager.views.form_builder.redirect", return_value="form-editor-response")
    @patch("corehq.apps.app_manager.views.form_builder.build_xform", return_value="<compiled-xform />")
    @patch("corehq.apps.app_manager.views.form_builder.Module.new_module")
    @patch("corehq.apps.app_manager.views.form_builder.Application.new_app")
    def test_explicit_save_creates_draft_and_redirects_to_editor(
        self,
        new_app,
        new_module,
        compile_xform,
        redirect_to_form,
        delete_preview,
        clear_app_cache,
        success_message,
    ):
        definition = parse_xlsform(_representative_xlsform(), "household.xlsx")
        form = MagicMock()
        form.get_unique_id.return_value = "form-1"
        module = MagicMock()
        module.new_form.return_value = form
        app = MagicMock()
        app.langs = ["en"]
        app.get_id = "app-1"
        app.add_module.return_value = module
        new_app.return_value = app
        new_module.return_value = module
        request = RequestFactory().post("/import/", {
            "app_name": "Field Research",
            "form_title": "Household Survey 2026",
            "module_unique_id": NEW_MODULE_VALUE,
            "module_name": "Baseline Surveys",
        })
        request.project = SimpleNamespace(secure_submissions=False)

        response = _save_draft(request, "test-domain", None, definition, "preview-1")

        self.assertEqual(response, "form-editor-response")
        new_app.assert_called_once_with("test-domain", "Field Research", lang="en")
        new_module.assert_called_once_with("Baseline Surveys", "en")
        app.add_module.assert_called_once_with(module)
        module.new_form.assert_called_once_with("Household Survey 2026", "en")
        compile_xform.assert_called_once_with(definition)
        self.assertEqual(form.source, "<compiled-xform />")
        self.assertEqual(app.langs, ["en", "fr"])
        app.save.assert_called_once_with()
        delete_preview.assert_called_once()
        clear_app_cache.assert_called_once_with(request, "test-domain")
        success_message.assert_called_once()
        redirect_to_form.assert_called_once_with("view_form", "test-domain", "app-1", "form-1")
