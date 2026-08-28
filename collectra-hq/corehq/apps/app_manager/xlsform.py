"""Parse SurveyCTO/XLSForm workbooks and compile them into CommCare XForms."""

import re
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook


XFORMS_NAMESPACE = "http://www.w3.org/2002/xforms"
XHTML_NAMESPACE = "http://www.w3.org/1999/xhtml"
JAVAROSA_NAMESPACE = "http://openrosa.org/javarosa"
OPENROSA_NAMESPACE = "http://openrosa.org/jr/xforms"
JRM_NAMESPACE = "http://dev.commcarehq.org/jr/xforms"
XSD_NAMESPACE = "http://www.w3.org/2001/XMLSchema"
VELLUM_NAMESPACE = "http://commcarehq.org/xforms/vellum"

MAX_XLSFORM_SIZE = 10 * 1024 * 1024
XML_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.-]*$")
REFERENCE_RE = re.compile(r"\$\{([^}]+)\}")
SELECTED_AT_REFERENCE_RE = re.compile(
    r"selected-at\s*\(\s*(\$\{([^}]+)\})\s*,\s*(\d+)\s*\)",
    re.IGNORECASE,
)
LOCALIZED_HEADER_RE = re.compile(
    r"^(label|hint|constraint_message|required_message)::(.+)$",
    re.IGNORECASE,
)

LANGUAGE_NAMES = {
    "arabic": "ar",
    "english": "en",
    "french": "fr",
    "hindi": "hi",
    "italian": "it",
    "portuguese": "pt",
    "spanish": "es",
    "swahili": "sw",
    "ukrainian": "uk",
}

TYPE_ALIASES = {
    "acknowledge": ("question", "boolean"),
    "audio": ("upload", "binary"),
    "barcode": ("question", "barcode"),
    "binary": ("upload", "binary"),
    "calculate": ("calculate", "string"),
    "date": ("question", "date"),
    "date time": ("question", "dateTime"),
    "datetime": ("question", "dateTime"),
    "decimal": ("question", "decimal"),
    "end": ("calculate", "dateTime"),
    "file": ("upload", "binary"),
    "geopoint": ("question", "geopoint"),
    "geoshape": ("question", "geoshape"),
    "geotrace": ("question", "geotrace"),
    "hidden": ("calculate", "string"),
    "image": ("upload", "binary"),
    "int": ("question", "int"),
    "integer": ("question", "int"),
    "note": ("note", "string"),
    "range": ("question", "int"),
    "start": ("calculate", "dateTime"),
    "string": ("question", "string"),
    "text": ("question", "string"),
    "time": ("question", "time"),
    "today": ("calculate", "date"),
    "video": ("upload", "binary"),
}

MEDIA_TYPES = {
    "audio": "audio/*",
    "file": "*/*",
    "image": "image/*",
    "video": "video/*",
}


class XlsFormError(Exception):
    """Raised when an uploaded workbook cannot be read as an XLSForm."""


@dataclass
class XlsFormIssue:
    level: str
    message: str
    sheet: str = "survey"
    row: int | None = None
    column: str = ""


@dataclass
class XlsFormChoice:
    row: int
    list_name: str
    name: str
    labels: dict[str, str]
    filters: dict[str, str] = field(default_factory=dict)

    @property
    def display_label(self):
        return next(iter(self.labels.values()), self.name)


@dataclass
class XlsFormRow:
    row: int
    raw_type: str
    kind: str
    data_type: str | None
    name: str
    labels: dict[str, str]
    hints: dict[str, str]
    path: list[str]
    list_name: str = ""
    required: str = ""
    required_messages: dict[str, str] = field(default_factory=dict)
    relevant: str = ""
    constraint: str = ""
    constraint_messages: dict[str, str] = field(default_factory=dict)
    calculation: str = ""
    default: str = ""
    appearance: str = ""
    choice_filter: str = ""

    @property
    def depth(self):
        return len(self.path)

    @property
    def is_select(self):
        return self.kind in {"select_one", "select_multiple"}

    @property
    def is_visible_question(self):
        return self.kind in {"question", "select_one", "select_multiple", "upload", "note"}

    @property
    def display_label(self):
        return next(iter(self.labels.values()), self.name)


@dataclass
class XlsFormDefinition:
    filename: str
    form_title: str
    form_id: str
    default_language: str
    version: str
    instance_name: str
    rows: list[XlsFormRow]
    choices: list[XlsFormChoice]
    issues: list[XlsFormIssue]

    @property
    def errors(self):
        return [issue for issue in self.issues if issue.level == "error"]

    @property
    def warnings(self):
        return [issue for issue in self.issues if issue.level == "warning"]

    @property
    def question_count(self):
        return sum(row.is_visible_question for row in self.rows)

    @property
    def languages(self):
        languages = {self.default_language}
        for row in self.rows:
            languages.update(row.labels)
            languages.update(row.hints)
            languages.update(row.constraint_messages)
            languages.update(row.required_messages)
        for choice in self.choices:
            languages.update(choice.labels)
        return sorted(language for language in languages if language)

    def to_dict(self):
        return asdict(self)

    @classmethod
    def from_dict(cls, data):
        return cls(
            filename=data["filename"],
            form_title=data["form_title"],
            form_id=data["form_id"],
            default_language=data["default_language"],
            version=data["version"],
            instance_name=data["instance_name"],
            rows=[XlsFormRow(**row) for row in data["rows"]],
            choices=[XlsFormChoice(**choice) for choice in data["choices"]],
            issues=[XlsFormIssue(**issue) for issue in data["issues"]],
        )


def _string(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    return str(value).strip()


def _normalize_header(value):
    return re.sub(r"\s+", "_", _string(value).lower())


def _normalize_type(value):
    return re.sub(r"[_\s]+", " ", _string(value).lower()).strip()


def _language_code(value, fallback="en"):
    value = _string(value)
    if not value:
        return fallback
    match = re.search(r"\(([A-Za-z][A-Za-z0-9-]*)\)\s*$", value)
    if match:
        return match.group(1).lower()
    lowered = value.lower()
    return LANGUAGE_NAMES.get(lowered, lowered.replace(" ", "-") or fallback)


def _sheet_records(sheet):
    values = sheet.iter_rows(values_only=True)
    first_row = next(values, None)
    if not first_row:
        return []
    headers = [_normalize_header(value) for value in first_row]
    if not any(headers):
        return []
    records = []
    for row_number, row in enumerate(values, 2):
        record = {
            header: _string(row[index]) if index < len(row) else ""
            for index, header in enumerate(headers)
            if header
        }
        if any(record.values()):
            records.append((row_number, record))
    return records


def _localized_values(record, base, default_language):
    values = {}
    plain_value = record.get(base, "")
    if plain_value:
        values[default_language] = plain_value
    prefix = f"{base}::"
    for header, value in record.items():
        if not value or not header.startswith(prefix):
            continue
        language = _language_code(header[len(prefix):], default_language)
        values[language] = value
    return values


def _parse_settings(sheet, filename):
    records = _sheet_records(sheet) if sheet is not None else []
    settings = records[0][1] if records else {}
    fallback_title = Path(filename).stem or "Untitled XLSForm"
    form_title = settings.get("form_title") or fallback_title
    form_id = settings.get("form_id") or re.sub(r"[^A-Za-z0-9_.-]+", "_", form_title).strip("_")
    if not form_id or not XML_NAME_RE.match(form_id):
        form_id = f"xlsform_{uuid.uuid4().hex[:12]}"
    return {
        "form_title": form_title,
        "form_id": form_id,
        "default_language": _language_code(settings.get("default_language"), "en"),
        "version": settings.get("version") or "1",
        "instance_name": settings.get("instance_name") or "",
    }


def _parse_choices(sheet, default_language, issues):
    if sheet is None:
        return []
    choices = []
    seen = set()
    standard_columns = {"list_name", "name"}
    for row_number, record in _sheet_records(sheet):
        list_name = record.get("list_name", "")
        name = record.get("name", "")
        if not list_name or not name:
            issues.append(XlsFormIssue(
                "error", "Every choice needs list_name and name.", "choices", row_number,
            ))
            continue
        key = (list_name, name)
        if key in seen:
            issues.append(XlsFormIssue(
                "error", f"Duplicate choice '{name}' in list '{list_name}'.", "choices", row_number, "name",
            ))
            continue
        seen.add(key)
        labels = _localized_values(record, "label", default_language)
        if not labels:
            labels = {default_language: name}
            issues.append(XlsFormIssue(
                "warning", f"Choice '{name}' has no label; its name will be displayed.",
                "choices", row_number, "label",
            ))
        filters = {
            header: value
            for header, value in record.items()
            if value
            and header not in standard_columns
            and header != "label"
            and not LOCALIZED_HEADER_RE.match(header)
        }
        choices.append(XlsFormChoice(row_number, list_name, name, labels, filters))
    return choices


def _parse_row_type(raw_type):
    raw_type = _string(raw_type)
    normalized = _normalize_type(raw_type)
    if normalized in {"begin group", "begin repeat"}:
        return ("begin_repeat" if normalized.endswith("repeat") else "begin_group", None, "")
    if normalized in {"end group", "end repeat"}:
        return ("end_repeat" if normalized.endswith("repeat") else "end_group", None, "")
    select_match = re.match(
        r"^select[_\s]+(one|multiple)(?:[_\s]+from[_\s]+file)?[_\s]+(.+)$",
        raw_type,
        re.IGNORECASE,
    )
    if select_match:
        kind = "select_one" if select_match.group(1) == "one" else "select_multiple"
        return kind, "string", select_match.group(2).strip()
    kind_and_type = TYPE_ALIASES.get(normalized)
    if kind_and_type:
        return (*kind_and_type, "")
    return "unsupported", None, ""


def _boolean_expression(value):
    normalized = _string(value).lower()
    if normalized in {"yes", "true", "true()", "1"}:
        return "true()"
    if normalized in {"no", "false", "false()", "0", ""}:
        return ""
    return _string(value)


def parse_xlsform(file_or_stream, filename=None):
    """Return a canonical XLSForm definition without modifying an application."""
    filename = filename or getattr(file_or_stream, "name", "uploaded.xlsx")
    try:
        workbook = load_workbook(file_or_stream, read_only=True, data_only=True)
    except Exception as exc:
        raise XlsFormError(f"The workbook could not be opened: {exc}") from exc

    try:
        sheets = {sheet.title.strip().lower(): sheet for sheet in workbook.worksheets}
        if "survey" not in sheets:
            raise XlsFormError("The workbook must contain a sheet named 'survey'.")

        settings = _parse_settings(sheets.get("settings"), filename)
        issues = []
        choices = _parse_choices(sheets.get("choices"), settings["default_language"], issues)
        choices_by_list = {}
        for choice in choices:
            choices_by_list.setdefault(choice.list_name, []).append(choice)

        rows = []
        group_stack = []
        seen_names = {}
        for row_number, record in _sheet_records(sheets["survey"]):
            raw_type = record.get("type", "")
            kind, data_type, list_name = _parse_row_type(raw_type)
            name = record.get("name", "")

            if not raw_type:
                issues.append(XlsFormIssue("error", "Question type is required.", "survey", row_number, "type"))
                continue
            if kind.startswith("end_"):
                expected = "begin_repeat" if kind == "end_repeat" else "begin_group"
                if not group_stack or group_stack[-1][0] != expected:
                    issues.append(XlsFormIssue(
                        "error", f"'{raw_type}' does not match the currently open group.",
                        "survey", row_number, "type",
                    ))
                else:
                    opened = group_stack.pop()
                    rows.append(XlsFormRow(
                        row_number, raw_type, kind, None, opened[1], {}, {},
                        [item[1] for item in group_stack],
                    ))
                continue

            if not name:
                issues.append(XlsFormIssue("error", "Question name is required.", "survey", row_number, "name"))
                continue
            if not XML_NAME_RE.match(name):
                issues.append(XlsFormIssue(
                    "error",
                    (
                        f"'{name}' is not a valid XML question name. "
                        "Use letters, numbers, dots, hyphens, or underscores."
                    ),
                    "survey", row_number, "name",
                ))
            if name in seen_names:
                issues.append(XlsFormIssue(
                    "error", f"Question name '{name}' was already used on row {seen_names[name]}.",
                    "survey", row_number, "name",
                ))
            else:
                seen_names[name] = row_number

            if kind == "unsupported":
                issues.append(XlsFormIssue(
                    "error", f"Question type '{raw_type}' is not supported yet.",
                    "survey", row_number, "type",
                ))

            labels = _localized_values(record, "label", settings["default_language"])
            hints = _localized_values(record, "hint", settings["default_language"])
            if kind not in {"calculate"} and not labels:
                labels = {settings["default_language"]: name}
                issues.append(XlsFormIssue(
                    "warning", f"Question '{name}' has no label; its name will be displayed.",
                    "survey", row_number, "label",
                ))

            if kind in {"select_one", "select_multiple"} and list_name not in choices_by_list:
                issues.append(XlsFormIssue(
                    "error", f"Choice list '{list_name}' was not found on the choices sheet.",
                    "survey", row_number, "type",
                ))

            path = [item[1] for item in group_stack]
            row = XlsFormRow(
                row=row_number,
                raw_type=raw_type,
                kind=kind,
                data_type=data_type,
                name=name,
                labels=labels,
                hints=hints,
                path=path,
                list_name=list_name,
                required=_boolean_expression(record.get("required")),
                required_messages=_localized_values(
                    record, "required_message", settings["default_language"],
                ),
                relevant=record.get("relevant", ""),
                constraint=record.get("constraint", ""),
                constraint_messages=_localized_values(
                    record, "constraint_message", settings["default_language"],
                ),
                calculation=record.get("calculation", ""),
                default=record.get("default", ""),
                appearance=record.get("appearance", ""),
                choice_filter=record.get("choice_filter", ""),
            )
            rows.append(row)
            if kind in {"begin_group", "begin_repeat"}:
                group_stack.append((kind, name, row_number))

        for kind, name, row_number in reversed(group_stack):
            closing_type = "end repeat" if kind == "begin_repeat" else "end group"
            issues.append(XlsFormIssue(
                "error", f"Group '{name}' is missing '{closing_type}'.", "survey", row_number, "type",
            ))

        if not rows:
            issues.append(XlsFormIssue("error", "The survey sheet has no question rows."))

        known_names = set(seen_names)
        for row in rows:
            expressions = {
                "relevant": row.relevant,
                "constraint": row.constraint,
                "calculation": row.calculation,
                "choice_filter": row.choice_filter,
            }
            for column, expression in expressions.items():
                for reference in REFERENCE_RE.findall(expression):
                    if reference not in known_names:
                        issues.append(XlsFormIssue(
                            "warning",
                            f"Expression references '${{{reference}}}', which is not a survey question.",
                            "survey", row.row, column,
                        ))

        return XlsFormDefinition(
            filename=Path(filename).name,
            form_title=settings["form_title"],
            form_id=settings["form_id"],
            default_language=settings["default_language"],
            version=settings["version"],
            instance_name=settings["instance_name"],
            rows=rows,
            choices=choices,
            issues=issues,
        )
    finally:
        workbook.close()


def _xforms_tag(name):
    return etree.QName(XFORMS_NAMESPACE, name)


def _xhtml_tag(name):
    return etree.QName(XHTML_NAMESPACE, name)


def _data_ref(row):
    return "/data/{}".format("/".join([*row.path, row.name]))


def _replace_references(expression, references):
    if not expression:
        return ""
    return REFERENCE_RE.sub(lambda match: references.get(match.group(1), match.group(0)), expression)


def _guard_selected_at_calculation(expression):
    """Keep direct list-item calculations empty until their source has enough values.

    SurveyCTO permits calculations such as ``selected-at(${gps}, 0)`` while the
    geopoint is empty. CommCare raises a runtime calculation error in that case,
    so direct extraction calculations need an explicit length guard.
    """
    if not expression:
        return ""
    match = SELECTED_AT_REFERENCE_RE.fullmatch(_string(expression))
    if not match:
        return expression
    reference = match.group(1)
    index = int(match.group(3))
    return "if(count-selected({}) > {}, {}, '')".format(
        reference,
        index,
        match.group(0),
    )


def _add_localized_text(translations, text_id, values, default_language):
    if not values:
        return
    fallback = next(iter(values.values()))
    for language in set(translations) | set(values):
        translations.setdefault(language, {})[text_id] = values.get(language, fallback)
    translations.setdefault(default_language, {})[text_id] = values.get(default_language, fallback)


def _append_label(control, text_id, hint_id=""):
    etree.SubElement(control, _xforms_tag("label"), {"ref": f"jr:itext('{text_id}')"})
    if hint_id:
        etree.SubElement(control, _xforms_tag("hint"), {"ref": f"jr:itext('{hint_id}')"})


def _build_choice_instance(model, row, choices, references):
    instance_id = f"choices_{row.name}"
    instance = etree.SubElement(model, _xforms_tag("instance"), {"id": instance_id})
    root = etree.SubElement(instance, "root")
    root.set("xmlns", "")
    for choice in choices:
        item = etree.SubElement(root, "item")
        etree.SubElement(item, "name").text = choice.name
        etree.SubElement(item, "label").text = "/".join([
            *row.path, f"{row.name}-{choice.name}-label",
        ])
        for key, value in choice.filters.items():
            if XML_NAME_RE.match(key):
                etree.SubElement(item, key).text = value
    filter_expression = _replace_references(row.choice_filter, references)
    nodeset = f"instance('{instance_id}')/root/item"
    if filter_expression:
        nodeset += f"[{filter_expression}]"
    return nodeset


def build_xform(definition):
    """Compile a validated canonical definition to a CommCare-compatible XForm."""
    if definition.errors:
        raise XlsFormError("The XLSForm has validation errors and cannot be compiled.")

    nsmap = {
        "h": XHTML_NAMESPACE,
        None: XFORMS_NAMESPACE,
        "jr": JAVAROSA_NAMESPACE,
        "orx": OPENROSA_NAMESPACE,
        "xsd": XSD_NAMESPACE,
        "vellum": VELLUM_NAMESPACE,
    }
    html = etree.Element(_xhtml_tag("html"), nsmap=nsmap)
    head = etree.SubElement(html, _xhtml_tag("head"))
    etree.SubElement(head, _xhtml_tag("title")).text = definition.form_title
    model = etree.SubElement(head, _xforms_tag("model"))
    instance = etree.SubElement(model, _xforms_tag("instance"))

    form_namespace = f"http://openrosa.org/formdesigner/{definition.form_id}"
    data = etree.SubElement(
        instance,
        etree.QName(form_namespace, "data"),
        nsmap={None: form_namespace, "jrm": JRM_NAMESPACE},
    )
    data.set("uiVersion", "1")
    data.set("version", definition.version)
    data.set("name", definition.form_title)

    references = {
        row.name: _data_ref(row)
        for row in definition.rows
        if not row.kind.startswith("end_")
    }
    choices_by_list = {}
    for choice in definition.choices:
        choices_by_list.setdefault(choice.list_name, []).append(choice)

    dynamic_choice_nodesets = {}
    for row in definition.rows:
        if row.is_select and row.choice_filter:
            dynamic_choice_nodesets[row.row] = _build_choice_instance(
                model,
                row,
                choices_by_list.get(row.list_name, []),
                references,
            )

    data_parents = {(): data}
    for row in definition.rows:
        if row.kind.startswith("end_"):
            continue
        parent = data_parents[tuple(row.path)]
        node = etree.SubElement(parent, etree.QName(form_namespace, row.name))
        if row.kind == "begin_repeat":
            node.set(etree.QName(JAVAROSA_NAMESPACE, "template"), "")
        if row.default:
            node.text = row.default
        if row.kind in {"begin_group", "begin_repeat"}:
            data_parents[tuple([*row.path, row.name])] = node

    translations = {language: {} for language in definition.languages}
    body = etree.SubElement(html, _xhtml_tag("body"))
    control_parents = {(): body}

    for row in definition.rows:
        if row.kind.startswith("end_"):
            continue
        ref = _data_ref(row)
        text_id = "/".join([*row.path, f"{row.name}-label"])
        hint_id = "/".join([*row.path, f"{row.name}-hint"]) if row.hints else ""
        constraint_id = "/".join([*row.path, f"{row.name}-constraintMsg"])
        required_id = "/".join([*row.path, f"{row.name}-requiredMsg"])
        _add_localized_text(translations, text_id, row.labels, definition.default_language)
        _add_localized_text(translations, hint_id, row.hints, definition.default_language)
        _add_localized_text(
            translations, constraint_id, row.constraint_messages, definition.default_language,
        )
        _add_localized_text(
            translations, required_id, row.required_messages, definition.default_language,
        )

        bind_attrs = {"nodeset": ref}
        if row.data_type:
            if row.data_type in {"string", "int", "boolean", "decimal", "date", "time", "dateTime"}:
                bind_attrs["type"] = f"xsd:{row.data_type}"
            else:
                bind_attrs["type"] = row.data_type
        if row.relevant:
            bind_attrs["relevant"] = _replace_references(row.relevant, references)
        if row.required:
            bind_attrs["required"] = _replace_references(row.required, references)
        if row.constraint:
            bind_attrs["constraint"] = _replace_references(row.constraint, references)
        calculation = row.calculation
        if row.kind == "calculate" and not calculation:
            calculation = {
                "start": "now()",
                "end": "now()",
                "today": "today()",
            }.get(_normalize_type(row.raw_type), "")
        if calculation:
            calculation = _guard_selected_at_calculation(calculation)
            bind_attrs["calculate"] = _replace_references(calculation, references)
        if row.kind == "note":
            bind_attrs["readonly"] = "true()"
        if row.constraint_messages:
            bind_attrs[etree.QName(JAVAROSA_NAMESPACE, "constraintMsg")] = f"jr:itext('{constraint_id}')"
        if row.required_messages:
            bind_attrs[etree.QName(JAVAROSA_NAMESPACE, "requiredMsg")] = f"jr:itext('{required_id}')"
        etree.SubElement(model, _xforms_tag("bind"), bind_attrs)

        parent = control_parents[tuple(row.path)]
        if row.kind in {"begin_group", "begin_repeat"}:
            group_attrs = {}
            if row.kind == "begin_group":
                group_attrs["ref"] = ref
            if row.appearance:
                group_attrs["appearance"] = row.appearance
            group = etree.SubElement(parent, _xforms_tag("group"), group_attrs)
            _append_label(group, text_id, hint_id)
            if row.kind == "begin_repeat":
                control_parent = etree.SubElement(group, _xforms_tag("repeat"), {"nodeset": ref})
            else:
                control_parent = group
            control_parents[tuple([*row.path, row.name])] = control_parent
            continue

        if row.kind == "calculate":
            continue

        control_attrs = {"ref": ref}
        if row.appearance:
            control_attrs["appearance"] = row.appearance
        if row.kind == "select_one":
            control = etree.SubElement(parent, _xforms_tag("select1"), control_attrs)
        elif row.kind == "select_multiple":
            control = etree.SubElement(parent, _xforms_tag("select"), control_attrs)
        elif row.kind == "upload":
            media_type = MEDIA_TYPES.get(_normalize_type(row.raw_type), "*/*")
            control_attrs["mediatype"] = media_type
            control = etree.SubElement(parent, _xforms_tag("upload"), control_attrs)
        elif row.kind == "note":
            control_attrs.setdefault("appearance", "minimal")
            control = etree.SubElement(parent, _xforms_tag("trigger"), control_attrs)
        else:
            control = etree.SubElement(parent, _xforms_tag("input"), control_attrs)
        _append_label(control, text_id, hint_id)

        if row.is_select:
            row_choices = choices_by_list.get(row.list_name, [])
            for choice in row_choices:
                choice_text_id = "/".join([
                    *row.path, f"{row.name}-{choice.name}-label",
                ])
                _add_localized_text(
                    translations, choice_text_id, choice.labels, definition.default_language,
                )
            if row.choice_filter:
                nodeset = dynamic_choice_nodesets[row.row]
                itemset = etree.SubElement(control, _xforms_tag("itemset"), {"nodeset": nodeset})
                etree.SubElement(itemset, _xforms_tag("label"), {"ref": "jr:itext(label)"})
                etree.SubElement(itemset, _xforms_tag("value"), {"ref": "name"})
            else:
                for choice in row_choices:
                    choice_text_id = "/".join([
                        *row.path, f"{row.name}-{choice.name}-label",
                    ])
                    item = etree.SubElement(control, _xforms_tag("item"))
                    etree.SubElement(
                        item, _xforms_tag("label"), {"ref": f"jr:itext('{choice_text_id}')"},
                    )
                    etree.SubElement(item, _xforms_tag("value")).text = choice.name

    itext = etree.SubElement(model, _xforms_tag("itext"))
    ordered_languages = [definition.default_language] + [
        language for language in definition.languages if language != definition.default_language
    ]
    for language in ordered_languages:
        attrs = {"lang": language}
        if language == definition.default_language:
            attrs["default"] = ""
        translation = etree.SubElement(itext, _xforms_tag("translation"), attrs)
        for text_id, value in translations.get(language, {}).items():
            text = etree.SubElement(translation, _xforms_tag("text"), {"id": text_id})
            etree.SubElement(text, _xforms_tag("value")).text = value

    return etree.tostring(
        html,
        encoding="unicode",
        pretty_print=True,
        xml_declaration=False,
    )
