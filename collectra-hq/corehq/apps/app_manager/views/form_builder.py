import re
from xml.sax.saxutils import escape

from django.contrib import messages
from django.shortcuts import render, redirect
from django.utils.translation import gettext as _

from corehq.apps.app_manager.models import Application, Module
from corehq.apps.app_manager.views.apps import clear_app_cache
from corehq.apps.app_manager.views.utils import generate_xmlns
from corehq.apps.domain.decorators import domain_admin_required
from corehq.apps.hqwebapp.decorators import use_bootstrap5


def _safe_xml_name(label, index):
    name = re.sub(r'[^A-Za-z0-9_]+', '_', label.strip().lower()).strip('_')
    if not name or name[0].isdigit():
        name = f'question_{index}'
    return name[:60]


def _build_xform(app_name, fields):
    xmlns = generate_xmlns()
    data_fields = []
    binds = []
    controls = []
    used = set()
    for index, field in enumerate(fields, 1):
        base = _safe_xml_name(field, index)
        name = base
        suffix = 2
        while name in used:
            name = f'{base}_{suffix}'
            suffix += 1
        used.add(name)
        question = escape(field)
        data_fields.append(f'          <{name} />')
        binds.append(f'          <bind nodeset="/data/{name}" type="xsd:string" />')
        controls.append(
            f'    <input ref="/data/{name}">\n'
            f'      <label>{question}</label>\n'
            f'    </input>'
        )

    return f'''<?xml version="1.0" encoding="UTF-8" ?>
<h:html xmlns:h="http://www.w3.org/1999/xhtml" xmlns:orx="http://openrosa.org/jr/xforms" xmlns="http://www.w3.org/2002/xforms" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:jr="http://openrosa.org/javarosa" xmlns:vellum="http://commcarehq.org/xforms/vellum">
  <h:head>
    <h:title>{escape(app_name)}</h:title>
    <model>
      <instance>
        <data xmlns:jrm="http://dev.commcarehq.org/jr/xforms" xmlns="http://openrosa.org/formdesigner/{xmlns}" uiVersion="1" version="1" name="{escape(app_name)}">
{chr(10).join(data_fields)}
        </data>
      </instance>
{chr(10).join(binds)}
    </model>
  </h:head>
  <h:body>
{chr(10).join(controls)}
  </h:body>
</h:html>
'''


def _builder_context(domain):
    return {'domain': domain}


def _headers_from_first_row(first_row):
    if not first_row:
        return []
    return [str(value).strip() for value in first_row if value not in (None, '')]


@use_bootstrap5
@domain_admin_required
def form_builder_choice(request, domain):
    return render(request, 'app_manager/form_builder_choice.html', _builder_context(domain))


@use_bootstrap5
@domain_admin_required
def excel_form_builder(request, domain):
    context = _builder_context(domain)
    if request.method == 'GET':
        return render(request, 'app_manager/excel_form_builder.html', context)

    upload = request.FILES.get('spreadsheet')
    app_name = (request.POST.get('app_name') or '').strip() or _('Untitled Excel Form')
    if not upload or not upload.name.lower().endswith(('.xlsx', '.xlsm')):
        messages.error(request, _('Please upload an Excel workbook (.xlsx or .xlsm).'))
        return render(request, 'app_manager/excel_form_builder.html', context)

    try:
        from openpyxl import load_workbook
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = _headers_from_first_row(first_row)
    except Exception as exc:
        messages.error(request, _('We could not read that spreadsheet: %s') % exc)
        return render(request, 'app_manager/excel_form_builder.html', context)

    if not headers:
        messages.error(request, _('The first row of the spreadsheet must contain question names.'))
        return render(request, 'app_manager/excel_form_builder.html', context)

    app = Application.new_app(domain, app_name, lang='en')
    module = Module.new_module(_('Survey'), 'en')
    app.add_module(module)
    form = app.new_form(0, app_name, 'en')
    form.source = _build_xform(app_name, headers)
    if getattr(request.project, 'secure_submissions', False):
        app.secure_submissions = True
    app.save()
    clear_app_cache(request, domain)
    messages.success(request, _('Your Excel form was created with %d questions.') % len(headers))
    return redirect('view_form', domain, app.id, form.get_unique_id())
